from openpyxl import load_workbook

def replace_text_in_excel(file_path, sheet_name):
    #Load the workbook and select the sheet
    print("Loading workbook...")
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        return
    ws = wb[sheet_name]
    print(f"Sheet '{sheet_name}' loaded successfully.")
    #₀₁₂₃₄₅₆₇₈₉

    source_name = ["'Cole Parmer'", "'Eldon James'", "'Phelps Elastomers'", "'Serfilco General CRG'",
                    "'Dupont'", "'TechnipFMC'", "'ARO Data'", "'BalSeal'", "'HiTechSeals'", "'ProcoProducts'", "'Serfilco Metals CRG'"]
    subscript_num = ["₁", "₂", "₃", "₄", "₅", "₆", "₇", "₈", "₉", "₁₀", "₁₁"]
    replace_total = [0] * len(source_name)

    cells_checked = 0
    replaced_count = 0

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str): 
                for i, name in enumerate(source_name): 
                    if name in cell.value:
                        new_value = cell.value.replace(name, subscript_num[i])
                        if new_value != cell.value:
                            cell.value = new_value
                            replace_total[i] += 1 
                            replaced_count += 1

            print(f"Checking cell: {cell.value}")
            cells_checked += 1

    
    wb.save(file_path)
    print(f"Replacements done ({replaced_count} occurrences) and saved in {file_path}")
    print(f"Cells checked: {cells_checked}")
    for i, count in enumerate(replace_total):
        print(f"{source_name[i]}: {count}")

file_path = "/Users/braxtonconley/Downloads/Proof of Concept.xlsx"  #Target File Path
sheet_name = "Chemicals"       #Target Sheet
replace_text_in_excel(file_path, sheet_name)
