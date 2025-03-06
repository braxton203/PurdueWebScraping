from openpyxl import load_workbook

def replace_text_in_excel(file_path, sheet_name):
    #Load the workbook and select the sheet
    print("Loading workbook...")
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        return
    ws = wb[sheet_name]
    print(f"Sheet '{sheet_name}' loaded successfully.")
    #₀₁₂₃₄₅₆₇₈₉
    #	1: A˱₂₃°   2: A˱₂₃ ͨ   3: A˱₂₁° ͨ
    #U+0368	   ͨ  872	Combining Latin Small Letter C
    #U+02F1	  ˱  753   Modifier Letter Low Left Arrowhead
    #U+00B0	  °	 176   0302 0260	&deg;	Degree sign
    original_temp = ['< 21C', '< 22C', '< 24C', '< 27C', '< 38C', '< 43C', '< 48C', '< 49C', '< 51C',
                    '< 52C', '< 55C', '< 60C', '< 66C', '< 70C', '< 71C', '< 77C', '< 80C', '< 82C',
                      '< 85C', '< 90C', '< 93C', '< 122C', '< 149C', '< 176C', '< 250C']
    new_temp = ['˱₂₁° ͨ', '˱₂₂° ͨ', '˱₂₄° ͨ', '˱₂₇° ͨ', '˱₃₈° ͨ', '˱₄₃° ͨ', '˱₄₈° ͨ', '˱₄₉° ͨ', '˱₅₁° ͨ', 
                 '˱₅₂° ͨ', '˱₅₅° ͨ', '˱₆₀° ͨ', '˱₆₆° ͨ', '˱₇₀° ͨ', '˱₇₁° ͨ', '˱₇₇° ͨ', '˱₈₀° ͨ', '˱₈₂° ͨ', 
                 '˱₈₅° ͨ', '˱₉₀° ͨ', '˱₉₃° ͨ', '˱₁₂₂° ͨ', '˱₁₄₉° ͨ', '˱₁₇₆° ͨ', '˱₂₅₀° ͨ']	
    replace_total = [0] * len(original_temp)

    cells_checked = 0
    replaced_count = 0

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str): 
                for i, name in enumerate(original_temp): 
                    if name in cell.value:
                        new_value = cell.value.replace(name, new_temp[i])
                        if new_value != cell.value:
                            cell.value = new_value
                            replace_total[i] += 1 
                            replaced_count += 1

            print(f"Checking cell: {cell.value}")
            cells_checked += 1

    
    wb.save(file_path)
    print(f"Replacements done ({replaced_count} occurrences) and saved in {file_path}")
    print(f"Cells checked: {cells_checked}")
    for i, count in enumerate(replace_total):
        print(f"{original_temp[i]}: {count}")

file_path = "/Users/braxtonconley/Downloads/Proof of Concept.xlsx"  #Target File Path
sheet_name = "Chemicals"       #Target Sheet
replace_text_in_excel(file_path, sheet_name)
